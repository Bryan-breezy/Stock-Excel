"""
Keeps a live Excel workbook mirrored to the database, per the PRD's
"Real-Time Sync Flow": every CRUD action on Product/Transaction/Supplier
appends or updates the matching row, and the Dashboard sheet holds
formulas only (no values written by Python).

This treats Postgres as the source of truth (see PRD section 16,
"Recommended architecture") — the workbook is a read-optimized mirror for
people who want a familiar spreadsheet view, not the system of record.

For a single-writer setup, direct openpyxl read/modify/save on each
change (as done here) is fine. If concurrent writers become an issue,
swap this for a queued/background sync (Celery task) instead of writing
inline in the request/response cycle.
"""

from openpyxl import Workbook, load_workbook
from django.conf import settings

PRODUCTS_HEADERS = ["ID", "SKU", "Product", "Category", "Qty", "Buy Price", "Sell Price", "Min Stock", "Supplier", "Last Updated"]
TRANSACTIONS_HEADERS = ["Date", "Product", "Type", "Qty", "Balance", "User", "Remarks"]
SUPPLIERS_HEADERS = ["Supplier", "Contact", "Phone", "Email"]


def _workbook_path():
    path = settings.EXCEL_WORKBOOK_PATH
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _get_or_create_workbook():
    path = _workbook_path()
    if path.exists():
        return load_workbook(path)

    wb = Workbook()
    products_ws = wb.active
    products_ws.title = "Products"
    products_ws.append(PRODUCTS_HEADERS)

    tx_ws = wb.create_sheet("Transactions")
    tx_ws.append(TRANSACTIONS_HEADERS)

    suppliers_ws = wb.create_sheet("Suppliers")
    suppliers_ws.append(SUPPLIERS_HEADERS)

    dashboard_ws = wb.create_sheet("Dashboard")
    _write_dashboard_formulas(dashboard_ws)

    return wb


def _write_dashboard_formulas(ws):
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A2"] = "Total inventory units"
    ws["B2"] = "=SUM(Products!E:E)"
    ws["A3"] = "Inventory value"
    ws["B3"] = "=SUMPRODUCT(Products!E2:E10000,Products!G2:G10000)"
    ws["A4"] = "Low stock items"
    ws["B4"] = '=COUNTIFS(Products!E:E,"<"&Products!H:H,Products!E:E,">0")'
    ws["A5"] = "Out of stock items"
    ws["B5"] = "=COUNTIF(Products!E:E,0)"


def _find_row_by_value(ws, column_letter, value, header_row=1):
    col_index = ord(column_letter.upper()) - ord("A")
    for row in ws.iter_rows(min_row=header_row + 1):
        if str(row[col_index].value) == str(value):
            return row[0].row
    return None


def upsert_product_row(product):
    wb = _get_or_create_workbook()
    ws = wb["Products"]
    row_data = [
        product.id,
        product.sku,
        product.name,
        product.category,
        product.quantity,
        float(product.buy_price),
        float(product.sell_price),
        product.minimum_stock,
        product.supplier.name if product.supplier else "",
        product.updated_at.strftime("%d/%m/%y %H:%M"),
    ]
    existing_row = _find_row_by_value(ws, "A", product.id)
    if existing_row:
        for col_index, value in enumerate(row_data, start=1):
            ws.cell(row=existing_row, column=col_index, value=value)
    else:
        ws.append(row_data)

    wb.save(_workbook_path())


def remove_product_row(product_id):
    wb = _get_or_create_workbook()
    ws = wb["Products"]
    row_number = _find_row_by_value(ws, "A", product_id)
    if row_number:
        ws.delete_rows(row_number)
    wb.save(_workbook_path())


def append_transaction_row(txn):
    wb = _get_or_create_workbook()
    ws = wb["Transactions"]
    ws.append([
        txn.created_at.strftime("%d/%m/%y"),
        txn.product.name,
        txn.type,
        txn.quantity if txn.type == "IN" else -txn.quantity,
        txn.balance_after,
        txn.created_by.username if txn.created_by else "",
        txn.remarks,
    ])
    wb.save(_workbook_path())


def upsert_supplier_row(supplier):
    wb = _get_or_create_workbook()
    ws = wb["Suppliers"]
    row_data = [supplier.name, supplier.contact_person, supplier.phone, supplier.email]
    existing_row = _find_row_by_value(ws, "A", supplier.name)
    if existing_row:
        for col_index, value in enumerate(row_data, start=1):
            ws.cell(row=existing_row, column=col_index, value=value)
    else:
        ws.append(row_data)
    wb.save(_workbook_path())
